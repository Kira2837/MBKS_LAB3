from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QVBoxLayout, QListWidget, QDialog, QWidget, QMessageBox, QTableWidgetItem, QCheckBox, QHBoxLayout
from PyQt5.QtCore import pyqtSignal
from UI_1 import Ui_MainWindow
import openpyxl
import os

class CheckBoxWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.checkBoxLayout = QHBoxLayout(self)
        self.checkBoxLayout.setAlignment(QtCore.Qt.AlignCenter)
        self.checkBox = QCheckBox()
        self.checkBox.setChecked(False)
        self.checkBoxLayout.addWidget(self.checkBox)
        self.setLayout(self.checkBoxLayout)

class DialogChange(QDialog):
    def __init__(self, title, ObjectChange):
        super().__init__()
        self.setWindowIcon(QtGui.QIcon('./pictures/update.png'))
        self.title = title
        self.ObjectChange = ObjectChange
        self.newName = ''
        
        self.setFixedSize(250, 200)
        self.setWindowTitle(title)
        layout = QVBoxLayout(self)
        self.name_label = QtWidgets.QLabel("Enter new name:")
        self.name_edit = QtWidgets.QLineEdit()
        self.name_edit.setPlaceholderText(f"New name {self.ObjectChange}")
        self.button_1 = QtWidgets.QPushButton("Submit")
        self.button_2 = QtWidgets.QPushButton("Cancel")
        self.name_label.setStyleSheet("font: 12pt \"Segoe UI\";")
        self.name_edit.setStyleSheet("font: 12pt \"Segoe UI\";")
        self.button_1.setStyleSheet("font: 12pt \"Segoe UI\";")
        self.button_2.setStyleSheet("font: 12pt \"Segoe UI\";")
        layout.addWidget(self.name_label)
        layout.addWidget(self.name_edit)
        layout.addWidget(self.button_1)
        layout.addWidget(self.button_2)
        
        self.button_1.clicked.connect(self.getNewName)
        self.button_2.clicked.connect(self.closeDiaglog)
    
    def getNewName(self):
        self.newName = self.name_edit.text()
        if (self.newName != ''):
            self.accept()
        else:
            QMessageBox.critical(self, "Error!", "Name cannot be empty!")
            
    def closeDiaglog(self):
        self.newName = ''
        self.accept()

class CreateWidget(QWidget):
    closed = pyqtSignal()
    
    def __init__(self):
        super().__init__()
        self.setFixedSize(250, 350)
        self.setWindowIcon(QtGui.QIcon("./pictures/add.png"))
        self.subject = None
        self.listObject = []
        self.setupWidget()
    
    def setupWidget(self):
        self.setWindowTitle("Create subject")
        layout = QVBoxLayout(self)
        self.label = QtWidgets.QLabel("Enter name subject:")
        self.name = QtWidgets.QLineEdit()
        self.name.setPlaceholderText("New subject")
        self.listWidget = QListWidget()
        self.object = QtWidgets.QLineEdit()
        self.object.setPlaceholderText("Enter object")
        self.button_1 = QtWidgets.QPushButton("Add")
        self.button_2 = QtWidgets.QPushButton("Remove")
        self.button_3 = QtWidgets.QPushButton("Submit")
        self.button_4 = QtWidgets.QPushButton("Cancel")
        layout.addWidget(self.label)
        layout.addWidget(self.name)
        layout.addWidget(self.listWidget)
        layout.addWidget(self.object)
        layout.addWidget(self.button_1)
        layout.addWidget(self.button_2)
        layout.addWidget(self.button_3)
        layout.addWidget(self.button_4)
        container = QWidget(self)
        container.setFixedSize(180, 340)
        container.setLayout(layout)
        container.move(40,10)
        
        self.button_1.clicked.connect(self.addObject)
        self.button_2.clicked.connect(self.removeObject)
        self.button_3.clicked.connect(self.closeAndSubmit)
        self.button_4.clicked.connect(self.closeWidget)
    
    def addObject(self):
        object = self.object.text().strip()
        if object and len(object) == 1 and object.isalnum() and object not in self.listObject:
            item = QtWidgets.QListWidgetItem(object)
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.listWidget.addItem(item)
            self.listObject.append(object)
            self.object.clear()
        elif object in self.listObject:
            self.ShowError("Error!", "Object has already been added!")
            self.object.clear()
        else:
            self.ShowError("Error!", "Invalid object name! (can only contain \nletter or number)!")
            self.object.clear()
    
    def removeObject(self):
        listItems = self.listWidget.selectedItems()
        if not listItems:
            return
        for item in listItems:
            self.listWidget.takeItem(self.listWidget.row(item))
            self.listObject.remove(item.text())
            if item is not None:
                item = None
    
    def closeAndSubmit(self):
        self.subject = self.name.text().strip()
        if self.subject and all(x.isalnum() for x in self.subject):
            pass
        else:
            self.ShowError("Error!", "Invalid subject name! (can only contain \nletters or numbers)!")
            return
        self.close()
        self.closed.emit()
    
    def closeWidget(self):
        self.subject = None
        self.listObjects = None
        self.close()
        self.closed.emit()
        
    def ShowError(self, Title, Message):
        error = QMessageBox()
        error.setWindowTitle(Title)
        error.setText(Message)
        error.setIcon(QMessageBox.Critical) 
        error.setStandardButtons(QMessageBox.Ok) 
        error.setDefaultButton(QMessageBox.Ok) 
        error.setWindowIcon(QtGui.QIcon('./pictures/notifi.png'))
        error.exec_()

class ManageAccessWidget(QWidget):
    closed = pyqtSignal()
    
    def __init__(self, listSubject, listObject, Mode):
        super().__init__()
        self.setFixedSize(400, 400)
        self.listSubject = listSubject
        self.listObject = listObject
        self.Mode = Mode
        self.listWidget = []
        self.setupWidget()
        
    def setupWidget(self):
        Layout = QHBoxLayout(self)
        for i in range(4): 
            listWidget = QListWidget()
            self.listWidget.append(listWidget)
            Layout.addWidget(listWidget)
        container = QWidget(self)
        container.setFixedSize(300, 300)
        container.setLayout(Layout)
        container.move(50,10)
        Layout_1 = QVBoxLayout(self)
        self.button_1 = QtWidgets.QPushButton("Submit")
        self.button_2 = QtWidgets.QPushButton("Cancel")
        self.button_1.setStyleSheet("font: 10pt \"Segoe UI\";")
        self.button_2.setStyleSheet("font: 10pt \"Segoe UI\";")
        Layout_1.addWidget(self.button_1)
        Layout_1.addWidget(self.button_2)
        container_1 = QWidget(self)
        container_1.setFixedSize(100, 100)
        container_1.setLayout(Layout_1)
        container_1.move(150,300)
        self.Mode_1 = QtWidgets.QRadioButton("None", self)
        self.Mode_1.move(50, 330)
        self.Mode_1.setChecked(True)
        self.Mode_2 = QtWidgets.QRadioButton(self)
        self.Mode_2.move(50, 360)
        self.addListSubject()
        self.addListObject()
        self.addListCheckBox()
        
        if self.Mode:
            self.setWindowTitle("Grant Access")
            self.setWindowIcon(QtGui.QIcon("./pictures/grant.png"))
            self.Mode_2.setText("Grant_all")
        else:
            self.setWindowTitle("Remove Access")
            self.setWindowIcon(QtGui.QIcon("./pictures/remove.png"))
            self.Mode_2.setText("Remove_all")
            
        self.button_1.clicked.connect(self.closeAndSubmit)
        self.button_2.clicked.connect(self.closeWidget)
        self.Mode_1.clicked.connect(self.setResetMode)
        self.Mode_2.toggled.connect(self.setModeAll)
        
    def addListSubject(self):
        for subject in self.listSubject:
            item = QtWidgets.QListWidgetItem(subject)
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            item.setSizeHint(QtCore.QSize(35, 35))
            self.listWidget[0].addItem(item)
    
    def addListObject(self):
        for object in self.listObject:
            item = QtWidgets.QListWidgetItem(object)
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            item.setSizeHint(QtCore.QSize(35, 35))
            self.listWidget[2].addItem(item)
    
    def addListCheckBox(self):
        for i in range(len(self.listSubject)):
            item = QtWidgets.QListWidgetItem(self.listWidget[1])
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            checkBox = CheckBoxWidget()
            item.setSizeHint(checkBox.sizeHint())
            self.listWidget[1].addItem(item)
            self.listWidget[1].setItemWidget(item, checkBox)
            
        for i in range(len(self.listObject)):
            item = QtWidgets.QListWidgetItem(self.listWidget[3])
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            checkBox = CheckBoxWidget()
            item.setSizeHint(checkBox.sizeHint())
            self.listWidget[3].addItem(item)
            self.listWidget[3].setItemWidget(item, checkBox)
    
    def getSelectedSubjects(self):
        self.ListSelectedSub = []
        for i in range(len(self.listSubject)):
            item = self.listWidget[1].item(i)
            checkBox = self.listWidget[1].itemWidget(item).layout().itemAt(0).widget()
            if checkBox.isChecked():
                self.ListSelectedSub.append(self.listSubject[i])
    
    # def getValueCheckBox(self):
    #     self.checkBoxBinVals = []
    #     for i in range(len(self.listObject)):
    #         item = self.listWidget[3].item(i)
    #         checkBox = self.listWidget[3].itemWidget(item).layout().itemAt(0).widget()
    #         self.checkBoxBinVals.append(int(checkBox.isChecked()))
    
    def getSelectedObjects(self):
        self.ListSelectedOb = []
        for i in range(len(self.listObject)):
            item = self.listWidget[3].item(i)
            checkBox = self.listWidget[3].itemWidget(item).layout().itemAt(0).widget()
            if checkBox.isChecked():
                self.ListSelectedOb.append(self.listObject[i])
    
    def setModeAll(self):
        radio_button = self.sender()      
        if radio_button.isChecked():
            for i in range(len(self.listObject)):
                item = self.listWidget[3].item(i)
                checkBox = self.listWidget[3].itemWidget(item).layout().itemAt(0).widget()
                checkBox.setChecked(True)
                checkBox.setEnabled(False)     
                checkBox.setStyleSheet('''QCheckBox::indicator:checked 
                {background-color: red;}''') 
    
    def setResetMode(self):
        radio_button = self.sender()
        
        if radio_button.isChecked():
            for i in range(len(self.listObject)):
                item = self.listWidget[3].item(i)
                checkBox = self.listWidget[3].itemWidget(item).layout().itemAt(0).widget()
                checkBox.setChecked(False)
                checkBox.setEnabled(True)     
                checkBox.setStyleSheet('') 
                
    def closeAndSubmit(self):
        self.getSelectedSubjects()
        self.getSelectedObjects()
        self.close()
        self.closed.emit()
    
    def closeWidget(self):
        self.ListSelectedSub = None
        self.ListSelectedOb = None
        self.close()
        self.closed.emit()
        
class MAIN_HANDLE(Ui_MainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setupUi(MainWindow)
        MainWindow.setWindowIcon(QtGui.QIcon('./pictures/admin.png'))
        MainWindow.setWindowTitle("Administrator")
        self.rowHeight = 50
        self.columnWidth = 50
        self.tableFilePath = "./table.xlsx"
        self.listObject = []
        self.listSubject = []
        self.checkBoxBinVals  = []
        self.listCheckBox = []
        
        self.getTableFromFile()
        self.setTableAndComboBox()

        self.Add_O.clicked.connect(self.addNewObject)
        self.Add_S.clicked.connect(self.addNewSubject)
        self.Delete_S.clicked.connect(self.delSubject)
        self.Delete_O.clicked.connect(self.delObject)
        self.Change_S.clicked.connect(self.changeSubject)
        self.Change_O.clicked.connect(self.changeObject)
        self.grant.clicked.connect(self.GrantAccess)
        self.remove.clicked.connect(self.RemoveAccess)
        self.create.clicked.connect(self.CreateFunc)
        self.Save.clicked.connect(self.saveTableToFile)
        self.Save.setDefault(True)

    def setTableAndComboBox(self):
        columns = len(self.listObject)
        rows = len(self.listSubject)
        
        self.tableWidget.setRowCount(rows)
        self.tableWidget.setColumnCount(columns)
        self.tableWidget.setHorizontalHeaderLabels(self.listObject)
        self.tableWidget.setVerticalHeaderLabels(self.listSubject)

        for row in range(rows):
            row_check_box = []
            self.tableWidget.setRowHeight(row, self.rowHeight)
            for column in range(columns):
                self.tableWidget.setColumnWidth(column, self.columnWidth)               
                check_box = CheckBoxWidget()
                check_box.checkBox.setChecked(self.checkBoxBinVals[row][column])
                self.tableWidget.setCellWidget(row, column, check_box)
                row_check_box.append(check_box)
                
            self.listCheckBox.append(row_check_box)

        for i in range(len(self.listSubject)):
            self.comboBox.addItem(self.listSubject[i])
        
        for i in range(len(self.listObject)):
            self.comboBox_2.addItem(self.listObject[i])
    
    def getTableFromFile(self):
        if (os.path.isfile(self.tableFilePath)):
            wb = openpyxl.load_workbook(self.tableFilePath)
            sheet = wb.active
            for subjectNum in range (2, sheet.max_row+1):
                cell = sheet.cell(subjectNum, 1)
                self.listSubject.append(cell.value)
            for objectNum in range (2, sheet.max_column+1):
                cell = sheet.cell(1, objectNum)
                self.listObject.append(cell.value)
            for subjectNum in range (2, sheet.max_row+1):
                rowBinVals = []
                for objectNum in range (2, sheet.max_column+1):
                    cell = sheet.cell(subjectNum, objectNum)
                    rowBinVals.append(cell.value)
                self.checkBoxBinVals.append(rowBinVals)
            wb.close()
        else:
            wb = openpyxl.Workbook()
            wb.save(self.tableFilePath)
            self.ShowStatus("Information!","file created no info in file yet!", True)
            wb.close()

    def saveTableToFile(self):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            rows = len(self.listSubject)
            columns = len(self.listObject)
            self.getValueCheckBox()

            for subjectNum in range (0, rows):
                cell = sheet.cell(subjectNum+2, 1)
                cell.value = self.listSubject[subjectNum]

            for objectNum in range (0, columns):
                cell = sheet.cell(1, objectNum+2)
                cell.value = self.listObject[objectNum]

            for subjectNum in range (0, rows):
                for objectNum in range (0, columns):
                    cell = sheet.cell(subjectNum+2, objectNum+2)
                    cell.value = self.checkBoxBinVals[subjectNum][objectNum]
            wb.save(self.tableFilePath)
            self.ShowStatus("Success!", "File saved successfully", True)
        except:
            self.ShowStatus("Error!", "File saved failed", False)
        finally:
            wb.close()
          
    def getValueCheckBox(self):
        self.checkBoxBinVals = []
        rows = len(self.listSubject)
        columns = len(self.listObject)
       
        for row in range(rows):
            rowVals = []
            for column in range(columns): 
                checkBox = self.tableWidget.cellWidget(row, column).layout().itemAt(0).widget()
                rowVals.append(int(checkBox.isChecked()))
            self.checkBoxBinVals.append(rowVals)
        
    def addNewSubject(self):
        nameSubject = self.subject.text()       
        if nameSubject and all(x.isalnum() for x in nameSubject):
            if nameSubject not in self.listSubject:
                row_position = self.tableWidget.rowCount()
                self.tableWidget.insertRow(row_position)
                columns = self.tableWidget.columnCount()
                self.tableWidget.setRowHeight(row_position, self.rowHeight)
                new_row_check_box = []
                rowVals = []
                for column in range(columns):
                    check_box = CheckBoxWidget()
                    self.tableWidget.setCellWidget(row_position, column, check_box)
                    new_row_check_box.append(check_box)
                    rowVals.append(int(check_box.checkBox.isChecked()))
                self.checkBoxBinVals.append(rowVals)    
                self.listCheckBox.append(new_row_check_box)
                self.tableWidget.setVerticalHeaderItem(row_position, QTableWidgetItem(nameSubject))
                self.listSubject.append(nameSubject)
                self.comboBox.addItem(nameSubject)
                self.ShowStatus("Success!", "Added new object!", True)
            else:
                self.ShowStatus("Error!", "Subject already exists!", False)
        else:
            self.ShowStatus("Error!", "Invalid subject name! (can only contain letters or numbers)!", False)
        self.subject.setText("")
        print(self.checkBoxBinVals)

    def addNewObject(self):
        nameObject = self.object.text()
        if nameObject and len(nameObject) == 1 and nameObject.isalnum():
            if nameObject not in self.listObject:
                column_position = self.tableWidget.columnCount()
                self.tableWidget.insertColumn(column_position)
                rows = self.tableWidget.rowCount()
                self.tableWidget.setColumnWidth(column_position, self.columnWidth)
                for row in range(rows):
                    check_box = CheckBoxWidget()
                    self.tableWidget.setCellWidget(row, column_position, check_box)
                    self.listCheckBox[row].append(check_box)
                    self.checkBoxBinVals[row].append(int(check_box.checkBox.isChecked()))
                self.tableWidget.setHorizontalHeaderItem(column_position, QTableWidgetItem(nameObject))
                self.listObject.append(nameObject)
                self.comboBox_2.addItem(nameObject)
                self.ShowStatus("Success!", "Added new subject!", True)
            else:
                self.ShowStatus("Error!", "Object already exists!", False)
        else:
            self.ShowStatus("Error!", "Invalid object name! (can only contain letter or number)!", False)
        self.object.setText("")

    def delSubject(self):
        nameSubject = self.comboBox.currentText()
        IdSubject = self.comboBox.currentIndex()
        if nameSubject in self.listSubject and IdSubject >= 0:
            self.comboBox.removeItem(IdSubject)
            self.tableWidget.removeRow(IdSubject)
            self.listSubject.remove(nameSubject)
            self.listCheckBox.remove(self.listCheckBox[IdSubject])
            self.ShowStatus("Success!", "Deleted subject", True)
   
    def delObject(self):
        nameObject = self.comboBox_2.currentText()
        IdObject = self.comboBox_2.currentIndex()
        if nameObject in self.listObject and IdObject >= 0:
            self.comboBox_2.removeItem(IdObject)
            self.tableWidget.removeColumn(IdObject)
            self.listObject.remove(nameObject)
            for i in range(len(self.listCheckBox)):
                self.listCheckBox[i].remove(self.listCheckBox[i][IdObject])
            self.ShowStatus("Success!", "Deleted object", True)

    def changeSubject(self):
        IdSubject = self.comboBox.currentIndex()
        if IdSubject == -1:
            return
        dialog = DialogChange("Edit subject", "subject")
        dialog.exec_()
        newName = dialog.newName

        if newName and all(x.isalnum() for x in newName):
            if (newName not in self.listSubject):
                self.comboBox.setItemText(IdSubject, newName)
                self.tableWidget.verticalHeaderItem(IdSubject).setText(newName)
                self.listSubject[IdSubject] = newName
                self.ShowStatus("Success!", "Updated new name Object!", True)
            else:
                self.ShowStatus("Error", "Update failed, new name already exists", False)
        elif newName:
            self.ShowStatus("Error!", "Invalid subject name! (can only contain letters or numbers)!", False)
            
    def changeObject(self):
        IdObject = self.comboBox_2.currentIndex()
        if IdObject == -1:
            return
        dialog = DialogChange("Edit object", "object")
        dialog.exec_()
        newName = dialog.newName
        
        if newName and len(newName) == 1 and newName.isalnum():
            if (newName and newName not in self.listObject):
                self.comboBox_2.setItemText(IdObject, newName)
                self.tableWidget.horizontalHeaderItem(IdObject).setText(newName)
                self.listObject[IdObject] = newName
                self.ShowStatus("Success!", "Updated new name Object!", True)
            else:
                self.ShowStatus("Error", "Update failed, new name already exists", False)
        elif newName:
            self.ShowStatus("Error!", "Invalid object name! (can only contain letter or number)!", False)
    
    def CreateFunc(self):
        self.widget_1 = CreateWidget()
        self.widget_1.closed.connect(self.onCreateWidgetClosed)
        self.widget_1.show()
    
    def onCreateWidgetClosed(self):
        newSubject = self.widget_1.subject
        if newSubject and newSubject not in self.listSubject:
            row_position = self.tableWidget.rowCount()
            self.tableWidget.insertRow(row_position)
            columns = self.tableWidget.columnCount()
            self.tableWidget.setRowHeight(row_position, self.rowHeight)
            new_row_check_box = []
            rowVals = []
            for column in range(columns):
                check_box = CheckBoxWidget()
                self.tableWidget.setCellWidget(row_position, column, check_box)
                new_row_check_box.append(check_box)
                rowVals.append(int(check_box.checkBox.isChecked()))
            self.checkBoxBinVals.append(rowVals) 
            self.listCheckBox.append(new_row_check_box)
            self.listSubject.append(newSubject)
            self.comboBox.addItem(newSubject)
            self.tableWidget.setVerticalHeaderItem(row_position, QTableWidgetItem(newSubject))
             
        newObjects = [object for object in self.widget_1.listObject if object not in self.listObject]
        rows = self.tableWidget.rowCount()
        if newObjects:   
            for object in newObjects:
                column_position = self.tableWidget.columnCount()
                self.tableWidget.insertColumn(column_position)
                self.tableWidget.setColumnWidth(column_position, self.columnWidth)
                for row in range(rows):
                    check_box = CheckBoxWidget()
                    self.tableWidget.setCellWidget(row, column_position, check_box)
                    self.listCheckBox[row].append(check_box)
                    self.checkBoxBinVals[row].append(int(check_box.checkBox.isChecked()))
                self.tableWidget.setHorizontalHeaderItem(column_position, QTableWidgetItem(object))
                self.listObject.append(object)
                self.comboBox_2.addItem(object)
        
        if newSubject:
            accessToObjects = self.widget_1.listObject
            index_S = self.listSubject.index(newSubject)
            for object in accessToObjects:
                index_O = self.listObject.index(object)
                self.checkBoxBinVals[index_S][index_O] = 1
            
            self.reloadTable()
                   
    def GrantAccess(self):
        self.widget = ManageAccessWidget(self.listSubject, self.listObject, True)
        self.widget.closed.connect(self.onGrantAccessWidgetClosed)
        self.widget.show() 
    
    def onGrantAccessWidgetClosed(self):
        if self.widget.ListSelectedSub and self.widget.ListSelectedOb:
            for subject in self.widget.ListSelectedSub:
                index_S = self.listSubject.index(subject)
                for object in self.widget.ListSelectedOb:
                    index_O = self.listObject.index(object)
                    self.checkBoxBinVals[index_S][index_O] = 1
        self.reloadTable()

    def RemoveAccess(self):
        self.widget = ManageAccessWidget(self.listSubject, self.listObject, False)
        self.widget.closed.connect(self.onRemoveAccessWidgetClosed)
        self.widget.show() 
    
    def onRemoveAccessWidgetClosed(self):
        if self.widget.ListSelectedSub and self.widget.ListSelectedOb:
            for subject in self.widget.ListSelectedSub:
                index_S = self.listSubject.index(subject)
                for object in self.widget.ListSelectedOb:
                    index_O = self.listObject.index(object)
                    self.checkBoxBinVals[index_S][index_O] = 0
        self.reloadTable()
        
    def reloadTable(self):
        columns = len(self.listObject)
        rows = len(self.listSubject)
        
        for row in range(rows):
            for column in range(columns):
                val = self.checkBoxBinVals[row][column]
                self.tableWidget.cellWidget(row, column).layout().itemAt(0).widget().setChecked(val)
        
    def ShowStatus(self, Title, Message, Status):
        error = QMessageBox()
        error.setWindowTitle(Title)
        error.setText(Message)
        if not Status:
            error.setIcon(QMessageBox.Critical) 
        else:
            error.setIcon(QMessageBox.Information) 
        error.setStandardButtons(QMessageBox.Ok) 
        error.setDefaultButton(QMessageBox.Ok) 
        error.setStyleSheet("font: 14pt \"MV Boli\";\n""color: rgb(255, 0, 0);")
        error.setWindowIcon(QtGui.QIcon('./pictures/notifi.png'))
        error.exec_()
        

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = MAIN_HANDLE()
    MainWindow.show()
    sys.exit(app.exec_())