from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QDialog, QMessageBox
from UI_1 import Ui_MainWindow
import openpyxl
import os

class Dialog(QDialog):
    def __init__(self, title, message):
        super().__init__()
        self.setWindowIcon(QtGui.QIcon('./pictures/message.png'))
    
        self.setFixedSize(200, 210)
        self.setWindowTitle(title)
        self.message = QtWidgets.QPlainTextEdit(message, self)
        self.message.setMaximumSize(200, 155)
        self.message.setReadOnly(True)
        self.button = QtWidgets.QPushButton("Ok", self)
        self.button.setGeometry(QtCore.QRect(70, 160, 70, 40))
        self.button.setStyleSheet("font: 12pt \"Segoe UI\";")
        
        self.button.clicked.connect(self.closeDiaglog)
    
    def closeDiaglog(self):
        self.accept()
        
class MAIN_HANDLE(Ui_MainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setupUi(MainWindow)
        MainWindow.setWindowIcon(QtGui.QIcon('./pictures/user.png'))
        self.tableFilePath = "./table.xlsx"
        self.listSubject = []
        self.listObject = []
        self.listAccess = []

        self.Btn.clicked.connect(self.getMessage)
        self.Btn.setDefault(True)
    
    def getMessage(self):
        self.getListNameOfTable()
        nameUser = self.name.text()
        if nameUser in self.listSubject:
            message = self.message.text()
            index_S = self.listSubject.index(nameUser)
            if message:
                for char in message:
                    if char in self.listObject:
                        index_O = self.listObject.index(char)
                        if not self.listAccess[index_S][index_O]:
                            message = message.replace(char, "")
                    else:
                        message = message.replace(char, "")
                if message:
                    dialog = Dialog(nameUser, message)
                    dialog.message.setStyleSheet("font: 14pt \"MV Boli\";")
                else:
                    dialog = Dialog("Error!", "Your message is empty after filter!")
                    dialog.message.setStyleSheet("font: 14pt \"MV Boli\";\n""color: rgb(255, 0, 0);")
                dialog.exec_()
            else:
                self.ShowStatus("Error!", "Message cannot be empty!", False)
        else:
            self.ShowStatus("Error", "Username does not exist", False)
            
    def getListNameOfTable(self):
        if (os.path.isfile(self.tableFilePath)):
            self.listSubject.clear()
            self.listObject.clear()
            self.listAccess.clear()
            wb = openpyxl.load_workbook(self.tableFilePath)
            sheet = wb.active
            for subjectNum in range (2, sheet.max_row+1):
                cell = sheet.cell(subjectNum, 1)
                self.listSubject.append(cell.value)
            for objectNum in range (2, sheet.max_column+1):
                cell = sheet.cell(1, objectNum)
                self.listObject.append(cell.value)
            for subjectNum in range (2, sheet.max_row+1):
                rowBinVals = []
                for objectNum in range (2, sheet.max_column+1):
                    cell = sheet.cell(subjectNum, objectNum)
                    rowBinVals.append(cell.value)
                self.listAccess.append(rowBinVals)
            wb.close()
        else:
            self.ShowStatus("Error!", "file created no info in file yet!", False)
    
    
    def ShowStatus(self, Title, Message, Status):
        error = QMessageBox()
        error.setWindowTitle(Title)
        error.setText(Message)
        if not Status:
            error.setIcon(QMessageBox.Critical) 
        else:
            error.setIcon(QMessageBox.Information) 
        error.setStandardButtons(QMessageBox.Ok) 
        error.setDefaultButton(QMessageBox.Ok) 
        error.setStyleSheet("font: 14pt \"MV Boli\";\n""color: rgb(255, 0, 0);")
        error.setWindowIcon(QtGui.QIcon('./pictures/notifi.png'))
        error.exec_()
        

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = MAIN_HANDLE()
    MainWindow.show()
    sys.exit(app.exec_())